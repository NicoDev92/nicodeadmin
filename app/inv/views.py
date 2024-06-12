from django.contrib.auth.models import User
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.db import IntegrityError
from django.http import JsonResponse
from django.utils import formats

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from openpyxl import load_workbook
from django.contrib.messages.views import SuccessMessageMixin


from django.views import generic
from django.urls import reverse_lazy

from .models import Categoria, SubCategoria, Marca, UnidadMedida, Producto, Precio
from .forms import CategoriaForm, SubCategoriaForm, MarcaForm, UMForm, ProductoForm, PrecioForm, ExcelUploadForm
from bases.views import SinAutorizacion


# Create your views here.

## CATEGORIAS
class CategoriaView(LoginRequiredMixin, SinAutorizacion, generic.ListView):
    model = Categoria
    permission_required = "inv.view_categoria"
    template_name = "inv/categoria_list.html"
    context_object_name = "obj"
    login_url = 'bases:login'
    
    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener el usuario de modificación para cada producto
        for categoria in queryset:
            if categoria.usuario_modifica:
                usuario_modifica = User.objects.filter(pk=categoria.usuario_modifica).first()
                categoria.usuario_modifica_nombre = usuario_modifica.username if usuario_modifica else None

        return queryset

class CategoriaNew(LoginRequiredMixin, SinAutorizacion, generic.CreateView):
    model = Categoria
    permission_required = "inv.add_categoria"
    template_name = "inv/categoria_form.html"
    context_object_name = "obj"
    form_class = CategoriaForm
    success_url = reverse_lazy("inv:categoria_list")
    login_url = "bases:login"

    def form_valid(self, form):
        form.instance.usuario_creador = self.request.user
        redirect_url = reverse_lazy("inv:categoria_list")
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('descripcion', 'La descripción ya existe. Por favor, ingrese una descripción diferente.')
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'errors': form.errors}, status=400)
            else:
                return super().form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = form.errors
            return JsonResponse({'errors': errors}, status=400)
        else:
            return super().form_invalid(form)

class CategoriaEdit(LoginRequiredMixin, SinAutorizacion, generic.UpdateView):
    model = Categoria
    permission_required = "inv.change_categoria"
    template_name = "inv/categoria_form.html"
    context_object_name = "obj"
    form_class = CategoriaForm
    success_url = reverse_lazy("inv:categoria_list")
    login_url = "bases:login"
    
    def form_valid(self, form):
        form.instance.usuario_modifica = self.request.user.id
        redirect_url = reverse_lazy("inv:categoria_list")
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('descripcion', 'La descripción ya existe. Por favor, ingrese una descripción diferente.')
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'errors': form.errors}, status=400)
            else:
                return super().form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = form.errors
            return JsonResponse({'errors': errors}, status=400)
        else:
            return super().form_invalid(form)

class CategoriaDel(LoginRequiredMixin, SinAutorizacion, generic.DeleteView):
    model=Categoria
    permission_required = "inv.del_categoria"
    template_name='inv/catalogos_del.html'
    context_object_name='obj'
    success_url=reverse_lazy("inv:categoria_list")

def categoria_inactivar(request, id):
    categoria = Categoria.objects.filter(pk=id).first()
    contexto = {}
    template_name = "inv/catalogos_del.html"
    
    if not categoria:
        return redirect("inv:categoria_list")
    
    if request.method == "GET":
        contexto = {'obj': categoria}
    
    if request.method == 'POST':
        # Alternar entre activo e inactivo
        categoria.usuario_modifica = request.user.id
        categoria.estado = not categoria.estado
        categoria.save()
        return redirect("inv:categoria_list")
    
    return render(request, template_name, contexto)

## SUBCATEGORIAS

class SubCategoriaView(LoginRequiredMixin, SinAutorizacion, generic.ListView):
    model = SubCategoria
    permission_required = "inv.view_subcategoria"
    template_name = "inv/subcategoria_list.html"
    context_object_name = "obj"
    login_url = 'bases:login'
    
    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener el usuario de modificación para cada producto
        for subcategoria in queryset:
            if subcategoria.usuario_modifica:
                usuario_modifica = User.objects.filter(pk=subcategoria.usuario_modifica).first()
                subcategoria.usuario_modifica_nombre = usuario_modifica.username if usuario_modifica else None

        return queryset

class SubCategoriaNew(LoginRequiredMixin, SinAutorizacion, generic.CreateView):
    model = SubCategoria
    permission_required = "inv.add_subcategoria"
    template_name = "inv/subcategoria_form.html"
    context_object_name = "obj"
    form_class = SubCategoriaForm
    success_url = reverse_lazy("inv:subcategoria_list")
    login_url = "bases:login"
    
    def form_valid(self, form):
        form.instance.usuario_creador = self.request.user
        redirect_url = reverse_lazy("inv:subcategoria_list")
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('descripcion', 'La descripción ya existe. Por favor, ingrese una descripción diferente.')
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'errors': form.errors}, status=400)
            else:
                return super().form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = form.errors
            return JsonResponse({'errors': errors}, status=400)
        else:
            return super().form_invalid(form)

class SubCategoriaEdit(LoginRequiredMixin, SinAutorizacion, generic.UpdateView):
    model = SubCategoria
    permission_required = "inv.change_subcategoria"
    template_name = "inv/subcategoria_form.html"
    context_object_name = "obj"
    form_class = SubCategoriaForm
    success_url = reverse_lazy("inv:subcategoria_list")
    login_url = "bases:login"
    
    def form_valid(self, form):
        form.instance.usuario_modifica = self.request.user.id
        redirect_url = reverse_lazy("inv:subcategoria_list")
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('descripcion', 'La descripción ya existe. Por favor, ingrese una descripción diferente.')
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'errors': form.errors}, status=400)
            else:
                return super().form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = form.errors
            return JsonResponse({'errors': errors}, status=400)
        else:
            return super().form_invalid(form)

class SubCategoriaDel(LoginRequiredMixin, SinAutorizacion, generic.DeleteView):
    model=SubCategoria
    template_name='inv/catalogos_del.html'
    context_object_name='obj'
    success_url=reverse_lazy("inv:subcategoria_list")

def subcategoria_inactivar(request, id):
    subcategoria = SubCategoria.objects.filter(pk=id).first()
    contexto = {}
    template_name = "inv/catalogos_del.html"
    
    if not subcategoria:
        return redirect("inv:subcategoria_list")
    
    if request.method == "GET":
        contexto = {'obj': subcategoria}
    
    if request.method == 'POST':
        # Alternar entre activo e inactivo
        subcategoria.usuario_modifica = request.user.id
        subcategoria.estado = not subcategoria.estado
        subcategoria.save()
        return redirect("inv:subcategoria_list")
    
    return render(request, template_name, contexto)

## MARCA

class MarcaView(LoginRequiredMixin, SinAutorizacion, generic.ListView):
    model = Marca
    permission_required = "inv.view_marca"
    template_name = "inv/marca_list.html"
    context_object_name = "obj"
    login_url = "bases:login"
    
    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener el usuario de modificación para cada producto
        for marca in queryset:
            if marca.usuario_modifica:
                usuario_modifica = User.objects.filter(pk=marca.usuario_modifica).first()
                marca.usuario_modifica_nombre = usuario_modifica.username if usuario_modifica else None

        return queryset

class MarcaNew(LoginRequiredMixin, SinAutorizacion, generic.CreateView):
    model = Marca
    permission_required = "inv.add_marca"
    template_name = "inv/marca_form.html"
    context_object_name = "obj"
    form_class = MarcaForm
    success_url = reverse_lazy("inv:marca_list")
    login_url = "bases:login"
    
    def form_valid(self, form):
        form.instance.usuario_creador = self.request.user
        redirect_url = reverse_lazy("inv:marca_list")
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('descripcion', 'La descripción ya existe. Por favor, ingrese una descripción diferente.')
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'errors': form.errors}, status=400)
            else:
                return super().form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = form.errors
            return JsonResponse({'errors': errors}, status=400)
        else:
            return super().form_invalid(form)

class MarcaEdit(LoginRequiredMixin, SinAutorizacion, generic.UpdateView):
    model = Marca
    permission_required = "inv.change_marca"
    template_name = "inv/marca_form.html"
    context_object_name = 'obj'
    form_class = MarcaForm
    success_url = reverse_lazy("inv:marca_list")
    success_message="Marca Creada"
    login_url = "bases:login"
    
    def form_valid(self, form):
        form.instance.usuario_modifica = self.request.user.id
        redirect_url = reverse_lazy("inv:marca_list")
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('descripcion', 'La descripción ya existe. Por favor, ingrese una descripción diferente.')
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'errors': form.errors}, status=400)
            else:
                return super().form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = form.errors
            return JsonResponse({'errors': errors}, status=400)
        else:
            return super().form_invalid(form)

def marca_inactivar(request, id):
    marca = Marca.objects.filter(pk=id).first()
    contexto = {}
    template_name = "inv/catalogos_del.html"
    
    if not marca:
        return redirect("inv:marca_list")
    
    if request.method == "GET":
        contexto = {'obj': marca}
    
    if request.method == 'POST':
        # Alternar entre activo e inactivo
        marca.usuario_modifica = request.user.id
        marca.estado = not marca.estado
        marca.save()
        return redirect("inv:marca_list")
    
    return render(request, template_name, contexto)

## UNIDAD DE MEDIDA

class UMView(LoginRequiredMixin, SinAutorizacion, generic.ListView):
    model = UnidadMedida
    permission_required = "inv.view_unidadmedida"
    template_name = "inv/um_list.html"
    context_object_name = "obj"
    login_url = 'bases:login'
    
    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener el usuario de modificación para cada producto
        for unidad_medida in queryset:
            if unidad_medida.usuario_modifica:
                usuario_modifica = User.objects.filter(pk=unidad_medida.usuario_modifica).first()
                unidad_medida.usuario_modifica_nombre = usuario_modifica.username if usuario_modifica else None

        return queryset

class UMNew(LoginRequiredMixin, SinAutorizacion, generic.CreateView):
    model=UnidadMedida
    permission_required = "inv.add_unidadmedida"
    template_name="inv/um_form.html"
    context_object_name = 'obj'
    form_class=UMForm
    success_url= reverse_lazy("inv:um_list")
    success_message="Unidad Medida Creada"
    login_url = 'bases:login'

    def form_valid(self, form):
        form.instance.usuario_creador = self.request.user
        redirect_url = reverse_lazy("inv:um_list")
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('descripcion', 'La descripción ya existe. Por favor, ingrese una descripción diferente.')
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'errors': form.errors}, status=400)
            else:
                return super().form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = form.errors
            return JsonResponse({'errors': errors}, status=400)
        else:
            return super().form_invalid(form)

class UMEdit(LoginRequiredMixin, SinAutorizacion, generic.UpdateView):
    model=UnidadMedida
    permission_required = "inv.change_unidadmedida"
    template_name="inv/um_form.html"
    context_object_name = 'obj'
    form_class=UMForm
    success_url= reverse_lazy("inv:um_list")
    success_message="Unidad Medida Editada"
    permission_required="inv.change_unidadmedida"

    def form_valid(self, form):
        form.instance.usuario_modifica = self.request.user.id
        redirect_url = reverse_lazy("inv:um_list")
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('descripcion', 'La descripción ya existe. Por favor, ingrese una descripción diferente.')
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'errors': form.errors}, status=400)
            else:
                return super().form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = form.errors
            return JsonResponse({'errors': errors}, status=400)
        else:
            return super().form_invalid(form)

def um_inactivar(request, id):
    um = UnidadMedida.objects.filter(pk=id).first()
    contexto = {}
    template_name = "inv/catalogos_del.html"
    
    if not um:
        return redirect("inv:um_list")
    
    if request.method == "GET":
        contexto = {'obj': um}
    
    if request.method == 'POST':
        # Alternar entre activo e inactivo
        um.usuario_modifica = request.user.id
        um.estado = not um.estado
        um.save()
        return redirect("inv:um_list")
    
    return render(request, template_name, contexto)

## PRODUCTOS

class ProductoView(LoginRequiredMixin, SinAutorizacion, generic.ListView):
    model = Producto
    permission_required = "inv.view_producto"
    template_name = "inv/prducto_list.html"
    context_object_name = "obj"
    login_url = 'bases:login'
    
    def get_queryset(self):
        queryset = super().get_queryset()

        # Obtener el usuario de modificación para cada producto
        for producto in queryset:
            if producto.usuario_modifica:
                usuario_modifica = User.objects.filter(pk=producto.usuario_modifica).first()
                producto.usuario_modifica_nombre = usuario_modifica.username if usuario_modifica else None


        return queryset

class ProductoNew(LoginRequiredMixin, SinAutorizacion, generic.CreateView):
    model = Producto
    permission_required = "inv.add_producto"
    template_name = "inv/producto_form.html"
    context_object_name = 'obj'
    form_class = ProductoForm
    success_url = reverse_lazy("inv:producto_list")
    success_message = "Producto Creado"

    def form_valid(self, form):
        form.instance.usuario_creador = self.request.user
        redirect_url =reverse_lazy("inv:producto_list")
        response = super().form_valid(form)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            redirect_url = self.success_url
            return JsonResponse({'redirect_url': redirect_url, 'success_message': self.success_message})
        return response
    
    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = form.errors
            return JsonResponse({'errors': errors}, status=400)
        else:
            return super().form_invalid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["marcas"] = Marca.objects.all()
        context["unidad_medidas"] = UnidadMedida.objects.all()
        context["categorias"] = Categoria.objects.all()
        context["subcategorias"] = SubCategoria.objects.all()
        return context

class ProductoEdit(LoginRequiredMixin, SinAutorizacion, SuccessMessageMixin, generic.UpdateView):
    model = Producto
    permission_required = "inv.change_producto"
    template_name = "inv/producto_form.html"
    context_object_name = 'obj'
    form_class = ProductoForm
    success_url = reverse_lazy("inv:producto_list")
    success_message = "Producto Editado"

    def form_valid(self, form):
        form.instance.usuario_modifica = self.request.user.id
        response = super().form_valid(form)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            redirect_url = self.success_url
            return JsonResponse({'redirect_url': redirect_url, 'success_message': self.success_message})
        return response

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = form.errors
            return JsonResponse({'errors': errors}, status=400)
        else:
            return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        pk = self.kwargs.get('pk')
        context = super(ProductoEdit, self).get_context_data(**kwargs)
        context["marcas"] = Marca.objects.all()
        context["unidad_medidas"] = UnidadMedida.objects.all()
        context["categorias"] = Categoria.objects.all()
        context["subcategorias"] = SubCategoria.objects.all()
        context["obj"] = Producto.objects.filter(pk=pk).first()
        obj = context.get('obj')
        if obj and obj.ultima_compra:
            context['ultima_compra_formatted'] = formats.date_format(obj.ultima_compra, "Y-m-d")
        return context

import traceback
from django.contrib import messages
from django.http import JsonResponse
from django.views import generic
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from .forms import ExcelUploadForm
from .models import Marca, UnidadMedida, SubCategoria, Producto

from django.http import JsonResponse

class ExcelUploadView(LoginRequiredMixin, SinAutorizacion, generic.FormView):
    form_class = ExcelUploadForm
    template_name = 'inv/product_list.html'
    success_url = reverse_lazy('inv:producto_list')
    permission_required = "inv.add_producto"

    def form_valid(self, form):
        response_data = {
            'success': True,
            'message': '',
            'file_message': '',
            'data': {
                'productos_creados': 0,
                'productos_modificados': 0,
                'productos_omitidos': 0,
                'detalles_productos': []
            }
        }

        try:
            file = form.cleaned_data['file']

            # Verificar que el archivo sea un archivo Excel
            if not file.name.endswith(('.xls', '.xlsx')):
                response_data['file_message'] = 'Formato de archivo no válido. Solo se permiten archivos .xls o .xlsx. (Excel)'
                return JsonResponse(response_data)

            wb = load_workbook(file)
            ws = wb.active

            for row in ws.iter_rows(values_only=True, min_row=2):
                try:
                    marca = Marca.objects.filter(descripcion=row[6].upper()).first() if row[6] else None
                    unidad_medida = UnidadMedida.objects.filter(descripcion=row[7].upper()).first() if row[7] else None
                    subcategoria = SubCategoria.objects.filter(descripcion=row[8].upper()).first() if row[8] else None

                    producto = Producto.objects.filter(codigo=row[0]).first()

                    if not producto:
                        producto = Producto(
                            codigo=row[0],
                            codigo_barra=row[1],
                            descripcion=row[2],
                            existencia=row[3],
                            ultima_compra=row[4],
                            precio_venta=row[5],
                            marca=marca,
                            unidad_medida=unidad_medida,
                            subcategoria=subcategoria,
                            usuario_creador=self.request.user
                        )
                        producto.save()
                        response_data['data']['productos_creados'] += 1
                        response_data['data']['detalles_productos'].append(f'Producto creado: {producto.descripcion}')
                    else:
                        producto.codigo_barra = row[1]
                        producto.descripcion = row[2]
                        producto.existencia = row[3]
                        producto.ultima_compra = row[4]
                        producto.precio_venta = row[5]
                        producto.marca = marca
                        producto.unidad_medida = unidad_medida
                        producto.subcategoria = subcategoria
                        producto.usuario_creador = self.request.user
                        producto.save()
                        response_data['data']['productos_modificados'] += 1
                        response_data['data']['detalles_productos'].append(f'Producto modificado: {producto.descripcion}')
                except Exception as e:
                    response_data['data']['productos_omitidos'] += 1
                    response_data['data']['detalles_productos'].append(f'Error al cargar producto: {str(e)}')

            response_data['message'] = f'{response_data["data"]["productos_creados"]} productos creados, {response_data["data"]["productos_modificados"]} productos modificados, {response_data["data"]["productos_omitidos"]} productos no cargados.'
        except Exception as e:
            response_data['success'] = False
            response_data['message'] = f"Error interno al cargar los productos: {str(e)}"
            traceback.print_exc()
        return JsonResponse(response_data)

    def form_invalid(self, form):
        response_data = {
            'success': False,
            'message': form.errors.as_json(),
            'file_message': '',
            'data': {
                'productos_creados': 0,
                'productos_modificados': 0,
                'productos_omitidos': 0,
                'detalles_productos': []
            }
        }
        return JsonResponse(response_data)


def producto_inactivar(request, id):
    prod = Producto.objects.filter(pk=id).first()
    contexto = {}
    template_name = "inv/catalogos_del.html"
    
    if not prod:
        return redirect("inv:producto_list")
    
    if request.method == "GET":
        contexto = {'obj': prod}
    
    if request.method == 'POST':
        # Alternar entre activo e inactivo
        prod.usuario_modifica = request.user.id
        prod.estado = not prod.estado
        prod.save()
        return redirect("inv:producto_list")
    
    return render(request, template_name, contexto)

##  PRECIOS

class PrecioNew(LoginRequiredMixin, SinAutorizacion, generic.CreateView):
    model = Precio
    permission_required = "inv.add_precio"
    form_class = PrecioForm
    template_name = "inv/precio_form.html"
    context_object_name = 'obj'
    success_url = reverse_lazy("inv:producto_list")
    success_message = "Pago parcial Nuevo"
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['producto_id'] = self.kwargs['id'] 
        return kwargs
    
    def form_valid(self, form):
        form.instance.usuario_creador = self.request.user
        form.instance.factura_id = self.kwargs['id'] 
        redirect_url = reverse_lazy("inv:producto_list")
        try:
            producto = Producto.objects.get(id=self.kwargs['id'])
            precios = Precio.objects.filter(producto=producto)
            

            if precios.exists():
                for precio in precios:
                    precio.estado = False
                    precio.save()

            producto.precio_venta = form.instance.precio_final
            producto.save()
            response = super().form_valid(form)
            return response
        
        except IntegrityError:
            """ form.add_error('descripcion', 'La descripción ya existe. Por favor, ingrese una descripción diferente.') """
            if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'errors': form.errors}, status=400)
            else:
                return super().form_invalid(form)

    def form_invalid(self, form):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            errors = form.errors
            return JsonResponse({'errors': errors}, status=400)
        else:
            return super().form_invalid(form)

